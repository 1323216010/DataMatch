import os
from configparser import ConfigParser
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
import xlrd

# 遍历当前文件夹，并获取所有文件的绝对路径
def get_file_paths(folder_path, include_subfolders=False):
    file_paths = []
    for foldername, subfolders, filenames in os.walk(folder_path):
        if not include_subfolders and foldername != folder_path:
            continue
        for filename in filenames:
            file_paths.append(os.path.join(foldername, filename))
    return file_paths

#控制台显示excel和csv文件读取过程
def read_file_with_progress(filename):
    _, ext = os.path.splitext(filename)  # 获取文件扩展名

    if ext == '.xlsx':
        # 使用openpyxl加载工作簿
        workbook = load_workbook(filename, read_only=True)
        sheet = workbook.active  # 获取活动工作表

        # 获取总行数，以便我们可以显示进度
        total_rows = sheet.max_row

        # 准备一个空的数据列表来保存数据
        data = []

        # 使用tqdm显示进度
        for row in tqdm(sheet.iter_rows(), total=total_rows):
            row_data = [cell.value for cell in row]
            data.append(row_data)

        # 转换为DataFrame并设置列名
        df = pd.DataFrame(data[1:], columns=data[0])

    elif ext == '.xls':
        # 使用xlrd加载工作簿
        workbook = xlrd.open_workbook(filename)
        sheet = workbook.sheet_by_index(0)

        # 获取总行数，以便我们可以显示进度
        total_rows = sheet.nrows

        # 使用tqdm显示进度
        data = [sheet.row_values(i) for i in tqdm(range(total_rows))]

        # 转换为DataFrame并设置列名
        df = pd.DataFrame(data[1:], columns=data[0])

    elif ext == '.csv':
        # 使用pandas的read_csv函数，但是分块读取
        chunk_size = 5000  # 可以根据需要调整块大小
        chunks = pd.read_csv(filename, chunksize=chunk_size)

        # 获取文件的总行数，以便我们可以显示进度
        total_rows = sum(1 for row in open(filename))

        data = []
        for chunk in tqdm(chunks, total=total_rows // chunk_size):
            data.append(chunk)

        # 合并所有块为一个DataFrame
        df = pd.concat(data, axis=0)

    else:
        raise ValueError(f"Unsupported file extension: {ext}")

    return df

def getConfig():
    config = ConfigParser()
    config.read(os.path.join(os.getcwd(), "config.ini"))
    vcmID = config['BuildData']['id']
    ActuatorSN = config['VCMData']['id']
    VCMDataPath = config['VCMData']['path']
    buildDataPath = config['BuildData']['path']
    precision = config['Rule']['precision']
    return vcmID, ActuatorSN, VCMDataPath, buildDataPath, precision

def printVersion():
    print("DataMatch version is 20230923a")
    print("What's news:")
    print("·Added the file read progress")
    print("Any question pls contact Pengcheng.yan@cowellchina.com")

#返回路径字符串中的文件格式
def check_path_type(path):
    if os.path.isdir(path):
        return 'Folder'
    elif os.path.isfile(path):
        file_name, file_extension = os.path.splitext(path)
        return file_extension.replace('.', '')
    else:
        return 'Not Found'
